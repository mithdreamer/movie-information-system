"""Excel dosyasından film adlarını okur ve sonuçları Excel'e yazar."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


HEADERS = [
    "Türkçe Film Adı",
    "İngilizce / Orijinal Ad",
    "Yapım Yılı",
    "Yönetmen",
    "Başrol Oyuncuları",
    "Tür",
    "Konu",
    "IMDb ID",
    "IMDb Puanı",
    "Durum",
]

RESULT_COLUMNS = [
    "input_name",
    "original_title",
    "year",
    "director",
    "actors",
    "genres",
    "overview",
    "imdb_id",
    "imdb_rating",
    "status",
]


def prepare_headers(worksheet):
    """Excel başlıklarını ve basit görünümü hazırlar."""
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for column_index, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill


def load_movie_names(file_path):
    """Excel'deki A kolonundan film adlarını okur."""
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    movie_names = []

    for row in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row, column=1).value

        if value is None:
            continue

        movie_name = str(value).strip()

        if movie_name:
            movie_names.append(movie_name)

    return movie_names


def write_movie_results(input_file_path, output_file_path, results):
    """Film sonuçlarını B-J kolonlarına yazar ve yeni Excel dosyası kaydeder."""
    workbook = load_workbook(input_file_path)
    worksheet = workbook.active

    prepare_headers(worksheet)

    for row_index, result in enumerate(results, start=2):
        for column_index, key in enumerate(RESULT_COLUMNS, start=1):
            worksheet.cell(row=row_index, column=column_index).value = result.get(key, "")

    output_path = Path(output_file_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
